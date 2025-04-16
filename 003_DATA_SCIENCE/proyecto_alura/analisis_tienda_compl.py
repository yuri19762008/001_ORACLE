import pandas as pd
import matplotlib.pyplot as plt
import os
import traceback
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

class AnalisisTienda:
    def __init__(self):
        self.resultados = {}
        os.makedirs('graficos', exist_ok=True)
        
    def cargar_datos(self, url):
        try:
            df = pd.read_csv(url)
            print(f"✅ Archivo cargado correctamente desde: {url}")
            return df
        except Exception as e:
            print(f"❌ Error al cargar datos desde {url}: {str(e)}")
            traceback.print_exc()
            return None
    
    def analizar_categoria(self, df, categoria):
        try:
            df_cat = df[df['Categoría del Producto'] == categoria]
            if df_cat.empty:
                print(f"⚠️ No hay datos para la categoría {categoria}")
                return {}
            resultados = {
                'Total Ventas': df_cat['Precio'].sum(),
                'Cantidad Productos': len(df_cat),
                'Precio Promedio': df_cat['Precio'].mean(),
                'Producto Más Vendido': df_cat['Producto'].value_counts().idxmax(),
                'Calificación Promedio': df_cat['Calificación'].mean(),
                'Costo Envío Promedio': df_cat['Costo de envío'].mean()
            }
            resultados['Top Productos'] = df_cat['Producto'].value_counts().head(3).index.tolist()
            coord_mas_repetida = df_cat.groupby(['lat', 'lon']).size().idxmax()
            resultados['Coordenada Más Repetida'] = f"({coord_mas_repetida[0]}, {coord_mas_repetida[1]})"
            return resultados
        except KeyError as e:
            print(f"❌ Error: Columna {e} no encontrada en los datos")
            return {}
        except Exception as e:
            print(f"❌ Error al analizar categoría {categoria}: {str(e)}")
            return {}
    
    def graficar_categoria(self, df, categoria, nombre_tienda):
        try:
            df_cat = df[df['Categoría del Producto'] == categoria]
            if df_cat.empty:
                print(f"⚠️ No hay datos para graficar la categoría {categoria}")
                return
            fig, axes = plt.subplots(2, 2, figsize=(15, 10))
            fig.suptitle(f'Análisis de {categoria} en {nombre_tienda}', fontsize=16)
            productos = df_cat['Producto'].value_counts().head(5)
            productos.plot(kind='bar', ax=axes[0, 0], color='blue')
            axes[0, 0].set_title(f'Top 5 Productos - {categoria}')
            axes[0, 0].set_ylabel('Cantidad Vendida')
            axes[0, 0].tick_params(axis='x', rotation=45)
            df_cat['Calificación'].value_counts().sort_index().plot(kind='bar', ax=axes[0, 1], color='green')
            axes[0, 1].set_title('Distribución de Calificaciones')
            axes[0, 1].set_xlabel('Calificación')
            axes[0, 1].set_ylabel('Cantidad')
            precio_por_producto = df_cat.groupby('Producto')['Precio'].mean().sort_values(ascending=False).head(5)
            precio_por_producto.plot(kind='barh', ax=axes[1, 0], color='red')
            axes[1, 0].set_title('Precio Promedio - Top 5 Productos')
            axes[1, 0].set_xlabel('Precio Promedio')
            df_cat['Método de pago'].value_counts().plot(kind='pie', ax=axes[1, 1], autopct='%1.1f%%')
            axes[1, 1].set_title('Métodos de Pago')
            plt.tight_layout(rect=[0, 0, 1, 0.95])
            nombre_imagen = f"graficos/{nombre_tienda}_{categoria.replace(' ', '_')}.png"
            plt.savefig(nombre_imagen)
            plt.close()
            print(f"📊 Gráfico de {categoria} generado en {nombre_tienda}")
        except Exception as e:
            print(f"❌ Error al graficar categoría {categoria}: {str(e)}")
            plt.close()
    
    def graficar_comparativa_categorias(self, resultados_tienda, nombre_tienda):
        try:
            categorias = []
            ingresos = []
            for cat, datos in resultados_tienda.items():
                if 'Total Ventas' in datos:
                    categorias.append(cat)
                    ingresos.append(datos['Total Ventas'])
            if not categorias:
                return
            plt.figure(figsize=(12, 6))
            plt.bar(categorias, ingresos, color='skyblue')
            plt.title(f'Ingresos por Categoría - {nombre_tienda}')
            plt.ylabel('Ingresos Totales')
            plt.xticks(rotation=45)
            plt.tight_layout()
            plt.savefig(f"graficos/{nombre_tienda}_categorias.png")
            plt.close()
        except Exception as e:
            print(f"❌ Error al generar comparativa de categorías: {str(e)}")
            plt.close()
    
    def analizar_tienda(self, nombre_tienda, url):
        try:
            print(f"\n{'='*50}")
            print(f"🏪 ANALIZANDO {nombre_tienda.upper()}")
            print(f"{'='*50}")
            df = self.cargar_datos(url)
            if df is None:
                return {}
            categorias = df['Categoría del Producto'].unique()
            print(f"📋 Categorías encontradas: {len(categorias)}")
            print(f"   {', '.join(categorias)}")
            resultados_tienda = {
                'ingresos_totales': df['Precio'].sum(),
                'calificacion_promedio': df['Calificación'].mean(),
                'productos_total': len(df),
                'categorias': {}
            }
            for categoria in categorias:
                print(f"\n📝 Analizando categoría: {categoria}")
                try:
                    resultados_categoria = self.analizar_categoria(df, categoria)
                    if resultados_categoria:
                        print(f"  • Producto más vendido: {resultados_categoria.get('Producto Más Vendido', 'N/A')}")
                        print(f"  • Total ventas: ${resultados_categoria.get('Total Ventas', 0):,.2f}")
                        print(f"  • Calificación promedio: {resultados_categoria.get('Calificación Promedio', 0):.2f}/5")
                    resultados_tienda['categorias'][categoria] = resultados_categoria
                    self.graficar_categoria(df, categoria, nombre_tienda)
                except Exception as e:
                    print(f"❌ Error procesando categoría {categoria}: {str(e)}")
                    continue
            self.graficar_comparativa_categorias(resultados_tienda['categorias'], nombre_tienda)
            self.resultados[nombre_tienda] = resultados_tienda
            return resultados_tienda
        except Exception as e:
            print(f"❌ Error al analizar tienda {nombre_tienda}: {str(e)}")
            traceback.print_exc()
            return {}
    
    def generar_recomendacion(self):
        try:
            if len(self.resultados) < 2:
                print("⚠️ Se necesitan al menos dos tiendas para generar recomendación")
                return
            print("\n\n📋 GENERANDO RECOMENDACIÓN FINAL...")
            metricas = {}
            for tienda, datos in self.resultados.items():
                metricas[tienda] = {
                    'ingresos': datos.get('ingresos_totales', 0),
                    'calificacion': datos.get('calificacion_promedio', 0),
                    'puntuacion': 0
                }
            for tienda, datos in metricas.items():
                ingresos_rel = datos['ingresos'] / max(m['ingresos'] for m in metricas.values()) if max(m['ingresos'] for m in metricas.values()) > 0 else 0
                calif_rel = datos['calificacion'] / 5
                datos['puntuacion'] = (ingresos_rel * 0.6) + (calif_rel * 0.4)
            tiendas_ordenadas = sorted(metricas.items(), key=lambda x: x[1]['puntuacion'])
            tienda_vender = tiendas_ordenadas[0][0]
            print("\n" + "="*60)
            print(f"📢 RECOMENDACIÓN FINAL: Vender {tienda_vender}")
            print("="*60)
            with open('recomendacion_final.txt', 'w') as f:
                f.write(f"RECOMENDACIÓN: Vender {tienda_vender}\n\n")
                for tienda, datos in tiendas_ordenadas:
                    f.write(f"{tienda}:\n")
                    f.write(f"  • Ingresos: ${datos['ingresos']:,.2f}\n")
                    f.write(f"  • Calificación: {datos['calificacion']:.2f}/5\n")
                    f.write(f"  • Puntuación: {datos['puntuacion']:.2f}\n\n")
            print("✅ Recomendación guardada en 'recomendacion_final.txt'")
        except Exception as e:
            print(f"❌ Error al generar recomendación: {str(e)}")
            traceback.print_exc()

    def exportar_excel(self, nombre_archivo="resumen_tiendas.xlsx"):
        try:
            print(f"\n📊 Exportando resultados a Excel: {nombre_archivo}")
            with pd.ExcelWriter(nombre_archivo) as writer:
                for tienda, datos in self.resultados.items():
                    resumen_tienda = pd.DataFrame({
                        'Ingresos Totales': [datos['ingresos_totales']],
                        'Calificación Promedio': [datos['calificacion_promedio']],
                        'Productos Totales': [datos['productos_total']]
                    })
                    resumen_tienda.to_excel(writer, sheet_name=f"{tienda}_Resumen", index=False)
                    categorias_data = []
                    for categoria, cat_datos in datos['categorias'].items():
                        if cat_datos:
                            row = {'Categoría': categoria}
                            row.update(cat_datos)
                            categorias_data.append(row)
                    if categorias_data:
                        df_categorias = pd.DataFrame(categorias_data)
                        df_categorias.to_excel(writer, sheet_name=f"{tienda}_Categorias", index=False)
            print(f"✅ Archivo Excel '{nombre_archivo}' generado correctamente")
            self.insertar_graficos_excel(nombre_archivo)
        except Exception as e:
            print(f"❌ Error al exportar a Excel: {str(e)}")
            traceback.print_exc()

    def insertar_graficos_excel(self, nombre_archivo):
        try:
            print(f"\n🖼️ Insertando gráficos en el archivo Excel: {nombre_archivo}")
            wb = load_workbook(nombre_archivo)
            for tienda in self.resultados.keys():
                hoja = f"{tienda}_Categorias"
                grafico_path = f"graficos/{tienda}_categorias.png"
                if hoja in wb.sheetnames and os.path.exists(grafico_path):
                    ws = wb[hoja]
                    img = XLImage(grafico_path)
                    img.width = 500
                    img.height = 320
                    ws.add_image(img, "J2")  # Inserta el gráfico en la celda J2
            wb.save(nombre_archivo.replace('.xlsx', '_con_graficos.xlsx'))
            print(f"✅ Gráficos insertados en '{nombre_archivo.replace('.xlsx', '_con_graficos.xlsx')}'")
        except Exception as e:
            print(f"❌ Error al insertar gráficos en Excel: {str(e)}")
            traceback.print_exc()

def main():
    try:
        urls = [
            "https://raw.githubusercontent.com/alura-es-cursos/challenge1-data-science-latam/refs/heads/main/base-de-datos-challenge1-latam/tienda_1%20.csv",
            "https://raw.githubusercontent.com/alura-es-cursos/challenge1-data-science-latam/refs/heads/main/base-de-datos-challenge1-latam/tienda_2.csv",
            "https://raw.githubusercontent.com/alura-es-cursos/challenge1-data-science-latam/refs/heads/main/base-de-datos-challenge1-latam/tienda_3.csv",
            "https://raw.githubusercontent.com/alura-es-cursos/challenge1-data-science-latam/refs/heads/main/base-de-datos-challenge1-latam/tienda_4.csv"
        ]
        nombres_tiendas = ["Tienda 1", "Tienda 2", "Tienda 3", "Tienda 4"]
        analizador = AnalisisTienda()
        for nombre, url in zip(nombres_tiendas, urls):
            analizador.analizar_tienda(nombre, url)
        analizador.generar_recomendacion()
        analizador.exportar_excel()
        print("\n✅ Análisis completado con éxito")
    except Exception as e:
        print(f"❌ Error en ejecución principal: {str(e)}")
        traceback.print_exc()

if __name__ == "__main__":
    main()
